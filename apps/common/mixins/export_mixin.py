"""
Export Mixin para ListView.
Agrega capacidad de exportar a CSV y Excel a cualquier ListView de Django.

Uso:
    class MiListView(ExportMixin, LoginRequiredMixin, ListView):
        model = MiModelo
        export_fields = ['field1', 'field2', 'field3']  # Campos a exportar
        export_filename = 'reporte'  # Nombre del archivo (sin extensión)

URLs soportadas:
    ?export=csv  - Descarga CSV
    ?export=excel - Descarga Excel
"""

import csv
from datetime import datetime

from django.http import HttpResponse


class ExportMixin:
    """
    Mixin que agrega exportación CSV/Excel a cualquier ListView.

    Attributes:
        export_fields: Lista de campos a exportar (opcional, usa todos si no se especifica)
        export_filename: Nombre base del archivo (default: 'reporte')
        export_related: Dict de campos relacionados a incluir
            Ejemplo: {'cliente__nombre': 'Cliente', 'moneda__codigo_iso': 'Moneda'}
    """

    export_fields = None
    export_filename = "reporte"
    export_related = {}

    def get_export_fields(self):
        """Retorna los campos a exportar."""
        if self.export_fields:
            return self.export_fields
        # Auto-detectar campos del modelo
        return [f.name for f in self.model._meta.fields if f.name != "id"]

    def get_export_data(self, queryset):
        """
        Procesa el queryset y retorna lista de dicts con los datos.
        Override este método para personalizar la exportación.
        """
        fields = self.get_export_fields()
        data = []

        for obj in queryset:
            row = {}
            for field in fields:
                # Manejar campos relacionados (ej: 'cliente__nombre')
                if "__" in field:
                    parts = field.split("__")
                    value = obj
                    for part in parts:
                        if value is None:
                            value = ""
                            break
                        value = getattr(value, part, "")
                    row[field] = str(value) if value else ""
                else:
                    value = getattr(obj, field, "")
                    # Formatear fechas
                    if hasattr(value, "strftime"):
                        row[field] = value.strftime("%Y-%m-%d %H:%M")
                    elif hasattr(value, "pk"):
                        row[field] = str(value)
                    else:
                        row[field] = value
            data.append(row)

        return data

    def get_export_headers(self):
        """
        Retorna headers legibles para las columnas.
        Override para personalizar nombres de columnas.
        """
        fields = self.get_export_fields()
        return [f.replace("_", " ").title() for f in fields]

    def export_csv(self, queryset):
        """Genera y retorna respuesta CSV."""
        data = self.get_export_data(queryset)
        headers = self.get_export_headers()

        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = (
            f'attachment; filename="{self.export_filename}_{datetime.now().strftime("%Y%m%d_%H%M")}.csv"'
        )

        # BOM para Excel en Windows
        response.write("\ufeff")

        writer = csv.writer(response)
        writer.writerow(headers)

        for row in data:
            writer.writerow([row.get(f, "") for f in self.get_export_fields()])

        return response

    def export_excel(self, queryset):
        """Genera y retorna respuesta Excel (XLSX)."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except ImportError:
            # Fallback a CSV si openpyxl no está instalado
            return self.export_csv(queryset)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.export_filename[:31]  # Excel limita a 31 chars

        # Headers con estilo
        headers = self.get_export_headers()
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        # Datos
        data = self.get_export_data(queryset)
        fields = self.get_export_fields()

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, field in enumerate(fields, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(field, ""))
                cell.border = thin_border

        # Auto-ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, len(data) + 2)
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(
                max_length + 2, 50
            )

        # Congregar headers
        ws.freeze_panes = "A2"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{self.export_filename}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        )

        wb.save(response)
        return response

    def export_pdf(self, queryset):
        """Genera y retorna respuesta PDF con WeasyPrint/Gotenberg."""
        import html
        import logging

        from django.utils import timezone

        from apps.common.services.pdf_renderer import PdfRendererService
        from core.api import get_user_active_agency

        local_logger = logging.getLogger(__name__)

        data = self.get_export_data(queryset)
        headers = self.get_export_headers()
        fields = self.get_export_fields()

        # Obtener datos de la agencia para personalización multi-tenant
        agencia = get_user_active_agency(self.request.user)
        agency_name = agencia.nombre if agencia else "TravelHub"

        # Formatear la fecha actual
        date_str = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %I:%M %p")

        # Título y metadatos
        title = f"Reporte de {self.export_filename.replace('_', ' ').title()}"

        # Construir cabeceras
        headers_html = "".join(f"<th>{html.escape(str(h))}</th>" for h in headers)

        # Construir filas
        rows_list = []
        for row in data:
            row_html = "<tr>"
            for field in fields:
                val = row.get(field, "")
                row_html += f"<td>{html.escape(str(val))}</td>"
            row_html += "</tr>"
            rows_list.append(row_html)
        rows_html = "".join(rows_list)

        # HTML completo para el PDF (Obsidian Emerald Theme)
        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{html.escape(title)}</title>
    <style>
        @page {{
            size: A4 landscape;
            margin: 15mm;
            @bottom-right {{
                content: "Página " counter(page) " de " counter(pages);
                font-family: Arial, sans-serif;
                font-size: 8pt;
                color: #6b7280;
            }}
            @bottom-left {{
                content: "{html.escape(agency_name)} - {html.escape(title)}";
                font-family: Arial, sans-serif;
                font-size: 8pt;
                color: #6b7280;
            }}
        }}
        body {{
            font-family: Arial, sans-serif;
            color: #1f2937;
            margin: 0;
            padding: 0;
            -webkit-print-color-adjust: exact;
        }}
        .header {{
            margin-bottom: 20px;
            border-bottom: 2px solid #047857;
            padding-bottom: 10px;
        }}
        .header h1 {{
            margin: 0;
            font-size: 20pt;
            color: #064e3b;
        }}
        .header .meta {{
            font-size: 9pt;
            color: #6b7280;
            margin-top: 5px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 8pt;
        }}
        th {{
            background-color: #047857;
            color: white;
            font-weight: bold;
            text-align: left;
            padding: 8px 10px;
            border: 1px solid #047857;
        }}
        td {{
            padding: 6px 10px;
            border: 1px solid #e5e7eb;
            word-break: break-all;
        }}
        tr:nth-child(even) td {{
            background-color: #f0fdf4; /* Obsidian emerald light tint */
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>{html.escape(title)}</h1>
        <div class="meta">
            <strong>Agencia:</strong> {html.escape(agency_name)} | 
            <strong>Fecha de Generación:</strong> {html.escape(date_str)} | 
            <strong>Total de Registros:</strong> {len(data)}
        </div>
    </div>
    <table>
        <thead>
            <tr>
                {headers_html}
            </tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
    </table>
</body>
</html>"""

        try:
            pdf_bytes = PdfRendererService.render_html_to_pdf(html_content)
            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename="{self.export_filename}_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
            )
            return response
        except Exception as e:
            local_logger.exception("Error al generar PDF de exportación")
            return HttpResponse(f"Error al generar reporte PDF: {str(e)}", status=500)

    def get(self, request, *args, **kwargs):
        """Intercepta GET para manejar exportación."""
        export_format = request.GET.get("export")

        if export_format in ("csv", "excel", "pdf"):
            # Obtener el queryset aplicando los mismos filtros que la vista normal
            queryset = self.get_queryset()

            # Limitar a 10,000 registros para evitar problemas de memoria
            if queryset.count() > 10000:
                queryset = queryset[:10000]

            if export_format == "csv":
                return self.export_csv(queryset)
            elif export_format == "excel":
                return self.export_excel(queryset)
            elif export_format == "pdf":
                return self.export_pdf(queryset)

        # Si no hay export, continuar con la vista normal
        return super().get(request, *args, **kwargs)

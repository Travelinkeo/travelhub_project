import io
import logging
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from django.apps import apps
from django.db.models import Q
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class AccountStatementService:
    """
    Servicio para generar el Estado de Cuenta / Relación de Cuenta Corriente del Cliente.
    Consolida:
      - Depósitos y Anticipos (Billetera / Saldo a Favor)
      - Ventas, Boletos y Servicios (Titular y Pasajeros Dependientes)
      - Consumos, Pagos y Deudas pendientes
      - Exportación a Excel (.xlsx) con formato ejecutivo corporativo
    """

    @classmethod
    def get_statement_data(
        cls,
        cliente,
        fecha_inicio: date | str | None = None,
        fecha_fin: date | str | None = None,
    ) -> dict:
        """
        Obtiene todos los datos financieros y transacciones del cliente y sus dependientes.
        """
        MovimientoSaldoCliente = apps.get_model("crm", "MovimientoSaldoCliente")
        Venta = apps.get_model("bookings", "Venta")

        # Parsear fechas si vienen como strings
        if isinstance(fecha_inicio, str) and fecha_inicio.strip():
            try:
                fecha_inicio = datetime.strptime(fecha_inicio.strip(), "%Y-%m-%d").date()
            except ValueError:
                fecha_inicio = None
        if isinstance(fecha_fin, str) and fecha_fin.strip():
            try:
                fecha_fin = datetime.strptime(fecha_fin.strip(), "%Y-%m-%d").date()
            except ValueError:
                fecha_fin = None

        # 1. Pasajeros dependientes vinculados
        pasajeros_dependientes = list(cliente.pasajeros.filter(is_deleted=False))
        pasajeros_ids = [p.pk for p in pasajeros_dependientes]

        # 2. Movimientos de saldo (Billetera)
        movs_qs = MovimientoSaldoCliente.all_objects.filter(
            cliente=cliente, is_deleted=False
        ).select_related("venta", "pago_venta", "moneda")

        if fecha_inicio:
            movs_qs = movs_qs.filter(creado__date__gte=fecha_inicio)
        if fecha_fin:
            movs_qs = movs_qs.filter(creado__date__lte=fecha_fin)

        movimientos = list(movs_qs.order_by("creado"))

        # 3. Ventas asociadas al cliente o a sus dependientes
        q_ventas = Q(cliente=cliente)
        if pasajeros_ids:
            q_ventas |= Q(pasajeros__id_pasajero__in=pasajeros_ids)

        ventas_qs = (
            Venta.all_objects.filter(q_ventas, is_deleted=False)
            .distinct()
            .select_related("moneda")
            .prefetch_related("pasajeros", "items_venta", "pagos_venta")
        )

        if fecha_inicio:
            ventas_qs = ventas_qs.filter(fecha_venta__date__gte=fecha_inicio)
        if fecha_fin:
            ventas_qs = ventas_qs.filter(fecha_venta__date__lte=fecha_fin)

        ventas = list(ventas_qs.order_by("fecha_venta"))

        # 4. Cálculos y Agregados
        total_anticipos = sum(
            (m.monto for m in movimientos if m.tipo_movimiento in ["DEP", "AJU"]),
            Decimal("0.00"),
        )
        total_consumido_saldo = sum(
            (m.monto for m in movimientos if m.tipo_movimiento in ["CON", "REE"]),
            Decimal("0.00"),
        )
        total_ventas_emitidas = sum((v.total_venta for v in ventas), Decimal("0.00"))
        total_ventas_pagadas = sum((v.monto_pagado for v in ventas), Decimal("0.00"))
        total_deuda_pendiente = sum((v.saldo_pendiente for v in ventas), Decimal("0.00"))

        saldo_a_favor_actual = cliente.saldo_a_favor

        # Balance neto entre anticipos recibidos y total de ventas generadas
        balance_neto = total_anticipos - total_ventas_emitidas

        return {
            "cliente": cliente,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
            "pasajeros_dependientes": pasajeros_dependientes,
            "movimientos": movimientos,
            "ventas": ventas,
            "total_anticipos": total_anticipos,
            "total_consumido_saldo": total_consumido_saldo,
            "total_ventas_emitidas": total_ventas_emitidas,
            "total_ventas_pagadas": total_ventas_pagadas,
            "total_deuda_pendiente": total_deuda_pendiente,
            "saldo_a_favor_actual": saldo_a_favor_actual,
            "balance_neto": balance_neto,
        }

    @classmethod
    def generate_excel_statement(
        cls,
        cliente,
        fecha_inicio: date | str | None = None,
        fecha_fin: date | str | None = None,
    ) -> io.BytesIO:
        """
        Genera un archivo Excel profesional (.xlsx) con la relación completa de cuenta corriente.
        """
        data = cls.get_statement_data(cliente, fecha_inicio, fecha_fin)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Estado de Cuenta"
        ws.views.sheetView[0].showGridLines = True

        # Paleta de Colores
        COLOR_PRIMARY = "1E293B"  # Slate 800
        COLOR_ACCENT = "4F46E5"  # Indigo 600
        COLOR_HEADER_BG = "0F172A"  # Dark Navy
        COLOR_TEXT_LIGHT = "FFFFFF"
        COLOR_SUCCESS_BG = "DCFCE7"
        COLOR_SUCCESS_TEXT = "166534"
        COLOR_DANGER_BG = "FEE2E2"
        COLOR_DANGER_TEXT = "991B1B"

        font_title = Font(name="Calibri", size=16, bold=True, color=COLOR_TEXT_LIGHT)
        font_subtitle = Font(name="Calibri", size=10, italic=True, color="94A3B8")
        font_header = Font(name="Calibri", size=11, bold=True, color=COLOR_TEXT_LIGHT)
        font_kpi_num = Font(name="Calibri", size=14, bold=True, color=COLOR_PRIMARY)
        font_kpi_lbl = Font(name="Calibri", size=9, bold=True, color="64748B")
        font_bold = Font(name="Calibri", size=10, bold=True)

        fill_header = PatternFill(
            start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid"
        )
        fill_accent = PatternFill(
            start_color=COLOR_ACCENT, end_color=COLOR_ACCENT, fill_type="solid"
        )
        fill_kpi = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        border_thin = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        # ── 1. BANNER DE ENCABEZADO ──────────────────────────────────────────
        ws.merge_cells("A1:G2")
        cell_banner = ws["A1"]
        agencia_nombre = cliente.agencia.nombre if cliente.agencia else "TravelHub ERP"
        cell_banner.value = f"{agencia_nombre.upper()} - ESTADO DE CUENTA Y RELACIÓN FINANCIERA"
        cell_banner.font = font_title
        cell_banner.fill = fill_header
        cell_banner.alignment = align_center

        ws["A3"].value = f"Generado el: {timezone.now().strftime('%d/%m/%Y %H:%M')} | Moneda: USD"
        ws["A3"].font = font_subtitle

        # ── 2. DATOS DEL CLIENTE ─────────────────────────────────────────────
        ws["A5"].value = "CLIENTE TITULAR:"
        ws["A5"].font = font_bold
        ws["B5"].value = cliente.get_nombre_completo()
        ws["B5"].font = font_bold

        ws["A6"].value = "DOCUMENTO:"
        ws["A6"].font = font_bold
        ws["B6"].value = cliente.cedula_identidad or cliente.numero_pasaporte or "S/D"

        ws["D5"].value = "PERÍODO:"
        ws["D5"].font = font_bold
        f_ini_str = (
            data["fecha_inicio"].strftime("%d/%m/%Y")
            if data["fecha_inicio"]
            else "Histórico Completo"
        )
        f_fin_str = data["fecha_fin"].strftime("%d/%m/%Y") if data["fecha_fin"] else "Presente"
        ws["E5"].value = f"{f_ini_str}  hasta  {f_fin_str}"

        ws["D6"].value = "DEPENDIENTES:"
        ws["D6"].font = font_bold
        nombres_deps = [p.get_nombre_completo() for p in data["pasajeros_dependientes"]]
        ws["E6"].value = ", ".join(nombres_deps) if nombres_deps else "Ninguno vinculado"

        # ── 3. TARJETAS RESUMEN / KPIS ───────────────────────────────────────
        # Caja 1: Total Anticipos
        ws.merge_cells("A8:B8")
        ws.merge_cells("A9:B9")
        ws["A8"].value = "TOTAL ANTICIPOS RECIBIDOS"
        ws["A8"].font = font_kpi_lbl
        ws["A8"].alignment = align_center
        ws["A8"].fill = fill_kpi
        ws["A9"].value = float(data["total_anticipos"])
        ws["A9"].number_format = "$#,##0.00"
        ws["A9"].font = font_kpi_num
        ws["A9"].alignment = align_center
        ws["A9"].fill = fill_kpi

        # Caja 2: Total Ventas / Emisiones
        ws.merge_cells("C8:D8")
        ws.merge_cells("C9:D9")
        ws["C8"].value = "TOTAL VENTAS / EMISIONES"
        ws["C8"].font = font_kpi_lbl
        ws["C8"].alignment = align_center
        ws["C8"].fill = fill_kpi
        ws["C9"].value = float(data["total_ventas_emitidas"])
        ws["C9"].number_format = "$#,##0.00"
        ws["C9"].font = font_kpi_num
        ws["C9"].alignment = align_center
        ws["C9"].fill = fill_kpi

        # Caja 3: Saldo Disponible o Deuda
        ws.merge_cells("E8:G8")
        ws.merge_cells("E9:G9")
        if data["saldo_a_favor_actual"] > Decimal("0.00"):
            ws["E8"].value = "SALDO A FAVOR DISPONIBLE"
            ws["E8"].fill = PatternFill(start_color=COLOR_SUCCESS_BG, fill_type="solid")
            ws["E8"].font = Font(name="Calibri", size=9, bold=True, color=COLOR_SUCCESS_TEXT)
            ws["E9"].value = float(data["saldo_a_favor_actual"])
            ws["E9"].font = Font(name="Calibri", size=14, bold=True, color=COLOR_SUCCESS_TEXT)
            ws["E9"].fill = PatternFill(start_color=COLOR_SUCCESS_BG, fill_type="solid")
        else:
            ws["E8"].value = "SALDO DEUDOR / DEUDA PENDIENTE"
            ws["E8"].fill = PatternFill(start_color=COLOR_DANGER_BG, fill_type="solid")
            ws["E8"].font = Font(name="Calibri", size=9, bold=True, color=COLOR_DANGER_TEXT)
            ws["E9"].value = float(data["total_deuda_pendiente"])
            ws["E9"].font = Font(name="Calibri", size=14, bold=True, color=COLOR_DANGER_TEXT)
            ws["E9"].fill = PatternFill(start_color=COLOR_DANGER_BG, fill_type="solid")

        ws["E8"].alignment = align_center
        ws["E9"].number_format = "$#,##0.00"
        ws["E9"].alignment = align_center

        for r in range(8, 10):
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = border_thin

        # ── 4. TABLA 1: DESGLOSE DE VENTAS Y SERVICIOS POR PASAJERO ──────────
        row = 11
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"].value = "1. RELACIÓN DE VENTAS Y EMISIONES (TITULAR Y DEPENDIENTES)"
        ws[f"A{row}"].font = font_header
        ws[f"A{row}"].fill = fill_accent
        ws[f"A{row}"].alignment = align_left

        row += 1
        headers_ventas = [
            "Fecha",
            "Localizador / PNR",
            "Pasajero / Beneficiario",
            "Concepto / Servicio",
            "Total Venta",
            "Pagado",
            "Saldo Pendiente",
        ]
        for col_num, h_text in enumerate(headers_ventas, 1):
            cell = ws.cell(row=row, column=col_num, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_right if col_num in [5, 6, 7] else align_left
            cell.border = border_thin

        for v in data["ventas"]:
            row += 1
            paxs = ", ".join([p.get_nombre_completo() for p in v.pasajeros.all()]) or (
                cliente.get_nombre_completo() if v.cliente == cliente else "Titular"
            )
            ws.cell(
                row=row, column=1, value=v.fecha_venta.strftime("%d/%m/%Y")
            ).alignment = align_left
            ws.cell(row=row, column=2, value=v.localizador or "S/L").alignment = align_center
            ws.cell(row=row, column=3, value=paxs).alignment = align_left
            ws.cell(
                row=row, column=4, value=v.descripcion_general or "Emisión de Boleto / Servicio"
            ).alignment = align_left

            c_tot = ws.cell(row=row, column=5, value=float(v.total_venta))
            c_tot.number_format = "$#,##0.00"
            c_tot.alignment = align_right

            c_pag = ws.cell(row=row, column=6, value=float(v.monto_pagado or 0))
            c_pag.number_format = "$#,##0.00"
            c_pag.alignment = align_right

            c_pen = ws.cell(row=row, column=7, value=float(v.saldo_pendiente or 0))
            c_pen.number_format = "$#,##0.00"
            c_pen.alignment = align_right

            for col_num in range(1, 8):
                ws.cell(row=row, column=col_num).border = border_thin

        if not data["ventas"]:
            row += 1
            ws.merge_cells(f"A{row}:G{row}")
            ws[f"A{row}"].value = "No se registran ventas para este período."
            ws[f"A{row}"].alignment = align_center
            ws[f"A{row}"].font = font_subtitle

        # Totales ventas
        row += 1
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"].value = "SUBTOTALES DE VENTAS:"
        ws[f"A{row}"].font = font_bold
        ws[f"A{row}"].alignment = align_right

        c_sub_tot = ws.cell(row=row, column=5, value=float(data["total_ventas_emitidas"]))
        c_sub_tot.font = font_bold
        c_sub_tot.number_format = "$#,##0.00"
        c_sub_tot.alignment = align_right

        c_sub_pag = ws.cell(row=row, column=6, value=float(data["total_ventas_pagadas"]))
        c_sub_pag.font = font_bold
        c_sub_pag.number_format = "$#,##0.00"
        c_sub_pag.alignment = align_right

        c_sub_pen = ws.cell(row=row, column=7, value=float(data["total_deuda_pendiente"]))
        c_sub_pen.font = font_bold
        c_sub_pen.number_format = "$#,##0.00"
        c_sub_pen.alignment = align_right

        # ── 5. TABLA 2: DESGLOSE DE ANTICIPOS Y MOVIMIENTOS DE SALDO ─────────
        row += 3
        ws.merge_cells(f"A{row}:G{row}")
        ws[f"A{row}"].value = "2. HISTORIAL DE ANTICIPOS Y MOVIMIENTOS EN BILLETERA"
        ws[f"A{row}"].font = font_header
        ws[f"A{row}"].fill = fill_accent
        ws[f"A{row}"].alignment = align_left

        row += 1
        headers_movs = [
            "Fecha",
            "Tipo de Movimiento",
            "Descripción / Motivo",
            "Método",
            "Referencia",
            "Monto",
            "Saldo Resultante",
        ]
        for col_num, h_text in enumerate(headers_movs, 1):
            cell = ws.cell(row=row, column=col_num, value=h_text)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_right if col_num in [6, 7] else align_left
            cell.border = border_thin

        for m in data["movimientos"]:
            row += 1
            ws.cell(
                row=row, column=1, value=m.creado.strftime("%d/%m/%Y %H:%M")
            ).alignment = align_left
            ws.cell(row=row, column=2, value=m.get_tipo_movimiento_display()).alignment = align_left
            ws.cell(row=row, column=3, value=m.descripcion or "").alignment = align_left
            ws.cell(row=row, column=4, value=m.metodo_pago_origen or "---").alignment = align_center
            ws.cell(row=row, column=5, value=m.referencia_bancaria or "---").alignment = align_left

            sign = "+" if m.tipo_movimiento in ["DEP", "AJU"] else "-"
            c_monto = ws.cell(row=row, column=6, value=float(m.monto))
            c_monto.number_format = f"{sign}$#,##0.00"
            c_monto.font = font_bold
            c_monto.alignment = align_right

            c_res = ws.cell(row=row, column=7, value=float(m.saldo_resultante))
            c_res.number_format = "$#,##0.00"
            c_res.alignment = align_right

            for col_num in range(1, 8):
                ws.cell(row=row, column=col_num).border = border_thin

        # Ajuste de ancho de columnas automático
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len and len(val_str) < 60:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

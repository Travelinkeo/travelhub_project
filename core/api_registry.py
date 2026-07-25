# core/api_registry.py
"""
Sistema de registro automático de APIs REST para modelos registrados en Django Admin.

Este módulo escanea los modelos registrados en admin.site y genera automáticamente
Serializers y ViewSets para exponerlos como APIs REST.
"""

import logging

from django.contrib import admin
from drf_spectacular.utils import extend_schema, extend_schema_view
from rest_framework import permissions, serializers, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from core.auth_helpers import InternalAPIAuthMixin

logger = logging.getLogger(__name__)

# Registry global para almacenar las APIs generadas
api_registry = {}


class AutoModelSerializer(serializers.ModelSerializer):
    """
    Serializer genérico que usa todos los campos del modelo.
    Campos sensibles (agencia, is_deleted, deleted_at, record_hash, estado) son read-only por defecto.
    """

    class Meta:
        """Configuración del modelo."""
        model = None
        fields = "__all__"
        read_only_fields = ("agencia", "is_deleted", "deleted_at", "record_hash", "estado")


@extend_schema_view(
    list=extend_schema(description="Listar todos los registros del modelo"),
    retrieve=extend_schema(description="Obtener un registro específico por ID"),
    create=extend_schema(description="Crear un nuevo registro"),
    update=extend_schema(description="Actualizar completamente un registro"),
    partial_update=extend_schema(description="Actualizar parcialmente un registro"),
    destroy=extend_schema(description="Eliminar un registro"),
)
class AutoModelViewSet(InternalAPIAuthMixin, viewsets.ModelViewSet):
    """
    ViewSet genérico para operaciones CRUD básicas.
    """

    serializer_class = None  # Se establece dinámicamente

    def get_permissions(self):
        """Método que obtiene permissions. Args: según implementación. Returns: datos solicitados."""
        if self.request.method in permissions.SAFE_METHODS:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), permissions.IsAdminUser()]

    def get_queryset(self):
        """Método que obtiene queryset. Args: según implementación. Returns: datos solicitados."""
        from core.middleware import get_current_agency

        model = self.serializer_class.Meta.model
        agency = get_current_agency()
        qs = model.objects.all()
        if agency and hasattr(model, "agencia"):
            qs = qs.filter(agencia=agency)
        elif agency and hasattr(model, "agency"):
            qs = qs.filter(agency=agency)
        return qs

    @extend_schema(description="Obtener el conteo total de registros")
    @action(detail=False, methods=["get"])
    def count(self, request):
        """
        Endpoint para obtener el conteo total de registros.
        """
        queryset = self.get_queryset()
        count = queryset.count()
        return Response({"count": count})

    @extend_schema(description="Exportar registros a un archivo Excel (.xlsx)")
    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        """
        Endpoint para exportar los registros del tenant actual a Excel.
        """
        from datetime import datetime

        import openpyxl
        from django.http import HttpResponse
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        queryset = self.get_queryset()

        # Limitar a 10,000 registros para evitar problemas de memoria
        if queryset.count() > 10000:
            queryset = queryset[:10000]

        model = self.serializer_class.Meta.model
        model_name = model.__name__

        fields = [
            f.name
            for f in model._meta.fields
            if f.name not in ("id", "agencia", "agency", "is_deleted", "deleted_at", "record_hash")
        ]
        headers = [f.replace("_", " ").title() for f in fields]

        data = []
        for obj in queryset:
            row = {}
            for field in fields:
                val = getattr(obj, field, "")
                if hasattr(val, "strftime"):
                    row[field] = val.strftime("%Y-%m-%d %H:%M")
                elif hasattr(val, "pk"):
                    row[field] = str(val)
                elif val is None:
                    row[field] = ""
                else:
                    row[field] = val
            data.append(row)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = model_name[:31]

        # Headers con estilo (Obsidian Emerald Theme)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        # Rellenar datos
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, field in enumerate(fields, 1):
                val = row_data.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border

        # Auto-ajustar columnas
        for col in range(1, len(headers) + 1):
            max_length = max(
                len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, len(data) + 2)
            )
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(
                max_length + 2, 50
            )

        ws.freeze_panes = "A2"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="reporte_{model_name.lower()}_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'
        )
        wb.save(response)
        return response

    @extend_schema(description="Exportar registros a un archivo PDF (.pdf)")
    @action(detail=False, methods=["get"])
    def export_pdf(self, request):
        """
        Endpoint para exportar los registros del tenant actual a PDF.
        """
        import html

        from django.http import HttpResponse
        from django.utils import timezone

        from apps.common.services.pdf_renderer import PdfRendererService
        from core.middleware import get_current_agency

        queryset = self.get_queryset()

        # Limitar a 10,000 registros para evitar problemas de memoria
        if queryset.count() > 10000:
            queryset = queryset[:10000]

        model = self.serializer_class.Meta.model
        model_name = model.__name__

        fields = [
            f.name
            for f in model._meta.fields
            if f.name not in ("id", "agencia", "agency", "is_deleted", "deleted_at", "record_hash")
        ]
        headers = [f.replace("_", " ").title() for f in fields]

        data = []
        for obj in queryset:
            row = {}
            for field in fields:
                val = getattr(obj, field, "")
                if hasattr(val, "strftime"):
                    row[field] = val.strftime("%Y-%m-%d %H:%M")
                elif hasattr(val, "pk"):
                    row[field] = str(val)
                elif val is None:
                    row[field] = ""
                else:
                    row[field] = val
            data.append(row)

        # Obtener datos de la agencia para personalización multi-tenant
        agency = get_current_agency()
        agency_name = agency.nombre if agency else "TravelHub"

        # Formatear la fecha actual
        date_str = timezone.localtime(timezone.now()).strftime("%d/%m/%Y %I:%M %p")

        # Título y metadatos
        title = f"Reporte de {model_name}"

        # Construir cabeceras
        headers_html = "".join(f"<th>{html.escape(str(h))}</th>" for h in headers)

        # Construir filas
        rows_list = []
        for row in data:
            row_html = "<tr>"
            for field in fields:
                val = row.get(field, "")
                row_html += f"<td>{html.escape(str(val))}</td>"
            row_html += "</tr>"
            rows_list.append(row_html)
        rows_html = "".join(rows_list)

        # HTML completo para el PDF (Obsidian Emerald Theme)
        html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{html.escape(title)}</title>
    <style>
        @page {{
            size: A4 landscape;
            margin: 15mm;
            @bottom-right {{
                content: "Página " counter(page) " de " counter(pages);
                font-family: Arial, sans-serif;
                font-size: 8pt;
                color: #6b7280;
            }}
            @bottom-left {{
                content: "{html.escape(agency_name)} - {html.escape(title)}";
                font-family: Arial, sans-serif;
                font-size: 8pt;
                color: #6b7280;
            }}
        }}
        body {{
            font-family: Arial, sans-serif;
            color: #1f2937;
            margin: 0;
            padding: 0;
            -webkit-print-color-adjust: exact;
        }}
        .header {{
            margin-bottom: 20px;
            border-bottom: 2px solid #047857;
            padding-bottom: 10px;
        }}
        .header h1 {{
            margin: 0;
            font-size: 20pt;
            color: #064e3b;
        }}
        .header .meta {{
            font-size: 9pt;
            color: #6b7280;
            margin-top: 5px;
        }}
        table {{
            width: 100%;
            border-collapse: collapse;
            font-size: 8pt;
        }}
        th {{
            background-color: #047857;
            color: white;
            font-weight: bold;
            text-align: left;
            padding: 8px 10px;
            border: 1px solid #047857;
        }}
        td {{
            padding: 6px 10px;
            border: 1px solid #e5e7eb;
            word-break: break-all;
        }}
        tr:nth-child(even) td {{
            background-color: #f0fdf4; /* Obsidian emerald light tint */
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>{html.escape(title)}</h1>
        <div class="meta">
            <strong>Agencia:</strong> {html.escape(agency_name)} |
            <strong>Fecha de Generación:</strong> {html.escape(date_str)} |
            <strong>Total de Registros:</strong> {len(data)}
        </div>
    </div>
    <table>
        <thead>
            <tr>
                {headers_html}
            </tr>
        </thead>
        <tbody>
            {rows_html}
        </tbody>
    </table>
</body>
</html>"""

        try:
            pdf_bytes = PdfRendererService.render_html_to_pdf(html_content)
            response = HttpResponse(pdf_bytes, content_type="application/pdf")
            response["Content-Disposition"] = (
                f'attachment; filename="reporte_{model_name.lower()}_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
            )
            return response
        except Exception as e:
            logger.exception("Error al generar PDF de exportación")
            return HttpResponse(f"Error al generar reporte PDF: {str(e)}", status=500)


def generate_api_for_model(model):
    """
    Genera Serializer y ViewSet para un modelo dado, prefiriendo uno existente si existe.
    """
    import core.serializers as core_serializers

    # Intentar buscar un serializer existente
    existing_serializer_name = f"{model.__name__}Serializer"
    SerializerClass = getattr(core_serializers, existing_serializer_name, None)

    if SerializerClass is None:
        FIELDS_MAP = {
            "AlquilerAutoReserva": [
                "venta",
                "proveedor",
                "ciudad_retiro",
                "ciudad_devolucion",
                "fecha_hora_retiro",
                "fecha_hora_devolucion",
                "categoria_auto",
                "compania_rentadora",
                "numero_confirmacion",
                "nombre_conductor",
                "incluye_seguro",
                "notas",
                "costo_neto",
                "precio_venta",
            ],
            "EventoServicio": [
                "venta",
                "proveedor",
                "nombre_evento",
                "fecha_evento",
                "ubicacion",
                "zona_asiento",
                "codigo_boleto_evento",
                "notas",
                "costo_neto",
                "precio_venta",
            ],
            "CircuitoTuristico": [
                "venta",
                "nombre_circuito",
                "dias_total",
                "fecha_inicio",
                "fecha_fin",
                "descripcion_general",
                "incluye",
                "no_incluye",
                "costo_neto_estimado",
                "precio_venta_estimado",
            ],
            "CircuitoDia": [
                "circuito",
                "dia_numero",
                "titulo",
                "descripcion",
                "ciudad",
                "alojamiento_previsto",
                "actividades_resumen",
            ],
            "PaqueteAereo": [
                "venta",
                "nombre_paquete",
                "incluye_vuelos",
                "incluye_hotel",
                "noches",
                "pasajeros",
                "resumen_componentes",
                "observaciones",
                "costo_neto_estimado",
                "precio_venta_estimado",
            ],
            "ServicioAdicionalDetalle": [
                "venta",
                "proveedor",
                "tipo_servicio",
                "descripcion",
                "codigo_referencia",
                "fecha_inicio",
                "fecha_fin",
                "nombre_pasajero",
                "notas",
                "costo_neto",
                "precio_venta",
            ],
            "Venta": [
                "localizador",
                "cliente",
                "fecha_venta",
                "descripcion_general",
                "moneda",
                "tasa_cambio_bcv",
                "subtotal",
                "impuestos",
                "total_venta",
                "monto_pagado",
                "saldo_pendiente",
                "estado",
                "tipo_venta",
                "canal_origen",
                "notas",
                "tiempo_limite_emision",
                "alerta_tl_disparada",
            ],
            "BoletoImportado": [
                "id_boleto_importado",
                "numero_boleto",
                "nombre_pasajero_completo",
                "total_boleto",
                "fecha_subida",
                "estado_parseo",
            ],
            "SegmentoVuelo": [
                "venta",
                "origen",
                "destino",
                "aerolinea",
                "numero_vuelo",
                "fecha_salida",
                "fecha_llegada",
                "clase_reserva",
                "cabina",
                "notas",
            ],
            "FeeVenta": [
                "venta",
                "tipo_fee",
                "descripcion",
                "monto",
                "moneda",
                "es_comision_agencia",
                "taxable",
            ],
            "PagoVenta": [
                "venta",
                "fecha_pago",
                "monto",
                "moneda",
                "metodo",
                "referencia",
                "confirmado",
                "aplica_igtf",
                "tasa_igtf",
                "monto_igtf",
                "notas",
            ],
        }

        fields = FIELDS_MAP.get(model.__name__, None)
        if fields is None:
            logger.warning(
                "No se generó API para %s: no hay field list definida. "
                "Agrega una entrada en FIELDS_MAP para exponer este modelo.",
                model.__name__,
            )
            return None, None

        # Crear Serializer dinámicamente
        serializer_name = f"{model.__name__}Serializer"
        serializer_attrs = {
            "Meta": type(
                "Meta",
                (),
                {
                    "model": model,
                    "fields": fields,
                    "read_only_fields": (
                        "agencia",
                        "is_deleted",
                        "deleted_at",
                        "record_hash",
                        "estado",
                    ),
                },
            )
        }
        SerializerClass = type(serializer_name, (AutoModelSerializer,), serializer_attrs)

    # Crear ViewSet dinámicamente
    viewset_name = f"{model.__name__}ViewSet"
    viewset_attrs = {
        "serializer_class": SerializerClass,
    }
    ViewSetClass = type(viewset_name, (AutoModelViewSet,), viewset_attrs)

    return SerializerClass, ViewSetClass


def register_auto_apis():
    """
    Escanea admin.site y registra APIs para todos los modelos registrados.
    """
    logger.debug(
        f"Modelos en admin.site._registry: {[model.__name__ for model in admin.site._registry.keys()]}"
    )
    logger.info("Iniciando registro automático de APIs...")
    WHITELIST = {
        "AlquilerAutoReserva",
        "EventoServicio",
        "CircuitoTuristico",
        "CircuitoDia",
        "PaqueteAereo",
        "ServicioAdicionalDetalle",
        "Venta",
        "BoletoImportado",
        "SegmentoVuelo",
        "FeeVenta",
        "PagoVenta",
    }
    for model, _admin_class in admin.site._registry.items():
        # Endurecimiento de Seguridad: Solo exponer modelos autorizados
        if not getattr(model, "api_expose", False) and model.__name__ not in WHITELIST:
            logger.debug(
                f"🛡️ Seguridad: Omitiendo auto-registro de API para el modelo {model.__name__} (sin api_expose=True)."
            )
            continue

        if model not in api_registry:
            try:
                serializer, viewset = generate_api_for_model(model)
                # Mapping for consistent basenames (matching tests)
                # Format: 'ModelName': ('singular-basename', 'plural-path')
                BASENAME_MAP = {
                    "AlquilerAutoReserva": ("alquiler-auto", "alquileres-autos"),
                    "EventoServicio": ("evento-servicio", "eventos-servicios"),
                    "CircuitoTuristico": ("circuito-turistico", "circuitos-turisticos"),
                    "CircuitoDia": ("circuito-dia", "circuitos-dias"),
                    "PaqueteAereo": ("paquete-aereo", "paquetes-aereos"),
                    "ServicioAdicionalDetalle": (
                        "servicio-adicional-detalle",
                        "servicios-adicionales",
                    ),
                    "Venta": ("venta", "ventas"),
                    "BoletoImportado": ("boletos-importados", "boletos-importados"),
                    "SegmentoVuelo": ("segmento-vuelo", "segmentos-vuelo"),
                    "FeeVenta": ("fee-venta", "fees-venta"),
                    "PagoVenta": ("pago-venta", "pagos-venta"),
                }

                if model.__name__ in BASENAME_MAP:
                    basename, path = BASENAME_MAP[model.__name__]
                else:
                    basename = model._meta.model_name
                    path = model._meta.model_name + "s"  # Default pluralization

                api_registry[model] = {
                    "serializer": serializer,
                    "viewset": viewset,
                    "basename": basename,
                    "path": path,
                }
                logger.info(f"API registrada para {model.__name__} con basename: {basename}")
            except Exception as e:
                logger.error(f"Error generando API para {model.__name__}: {e}")
    logger.info(f"Total APIs en registry: {len(api_registry)}")
    logger.debug(f"Basenames registrados: {[api['basename'] for api in api_registry.values()]}")


def get_registered_apis():
    """
    Retorna un diccionario con todas las APIs registradas.
    """
    return api_registry

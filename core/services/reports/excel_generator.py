import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO

class ExcelGenerator:
    """
    Service for generating Excel files from datasets.
    """

    def __init__(self, headers, data, sheet_name="Reporte"):
        """
        :param headers: List of strings for column headers.
        :param data: List of lists/tuples executing the data rows.
        :param sheet_name: Name of the worksheet.
        """
        self.headers = headers
        self.data = data
        self.sheet_name = sheet_name
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = self.sheet_name

    def generate(self):
        """
        Generates the Excel file in memory and returns the BytesIO object.
        """
        self._write_headers()
        self._write_data()
        self._style_worksheet()
        
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output

    def _write_headers(self):
        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo-600
        center_alignment = Alignment(horizontal="center", vertical="center")

        for col_num, header_title in enumerate(self.headers, 1):
            cell = self.ws.cell(row=1, column=col_num, value=header_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment

    def _write_data(self):
        for row_num, row_data in enumerate(self.data, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = self.ws.cell(row=row_num, column=col_num, value=cell_value)
                
                # Format dates and numbers
                if isinstance(cell_value, datetime):
                    cell.number_format = 'DD/MM/YYYY'
                elif isinstance(cell_value, (int, float)):
                    # Check if it looks like currency (heuristic or explicit type check if possible)
                    pass 
                
                # Alternate row colors for readability
                if row_num % 2 == 0:
                    cell.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid") # Gray-50

    def _style_worksheet(self):
        # Auto-adjust column widths
        for col_num, column_cells in enumerate(self.ws.columns, 1):
            max_length = 0
            column = get_column_letter(col_num)
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            self.ws.column_dimensions[column].width = adjusted_width

        # Add thin border to all cells
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        for row in self.ws.iter_rows():
            for cell in row:
                cell.border = thin_border

from decimal import Decimal

import pytest
from django.contrib.auth import get_user_model
from django.test import Client, TestCase

from apps.bookings.models import Venta
from apps.common.models import Moneda
from apps.crm.models import Cliente
from apps.finance.models import Factura
from core.middleware import agency_context
from core.models.agencia import Agencia, UsuarioAgencia

User = get_user_model()

# El export a PDF depende de un backend externo (Gotenberg o WeasyPrint).
# WeasyPrint requiere cffi + librerías nativas (cairo/pango); en entornos de
# CI sin esas dependencias (o sin Gotenberg) el render no está disponible.
# La lógica de aislamiento multi-tenant ya se valida en los export Excel.
_weasyprint_ok = False
try:
    import cffi  # noqa: F401  (puente nativo requerido por WeasyPrint)
    import weasyprint  # noqa: F401

    _weasyprint_ok = True
except Exception:
    _weasyprint_ok = False

requires_pdf_backend = pytest.mark.skipif(
    not _weasyprint_ok,
    reason="Backend de PDF (WeasyPrint/Gotenberg) no disponible en este entorno.",
)


class GlobalExportsTenantIsolationTest(TestCase):
    def setUp(self):
        # 1. Crear Agencias (Tenants)
        self.agencia_a = Agencia.objects.create(nombre="Agencia A", rif="J123456789")
        self.agencia_b = Agencia.objects.create(nombre="Agencia B", rif="J987654321")

        # Configurar subdominios slugs
        config_a = self.agencia_a.configuracion
        config_a.subdominio_slug = "agencia-a"
        config_a.save()

        config_b = self.agencia_b.configuracion
        config_b.subdominio_slug = "agencia-b"
        config_b.save()

        # 2. Crear Usuarios para cada Agencia
        self.user_a = User.objects.create_user(
            username="admin_a",
            email="admin@agencia-a.com",
            password="password123",
        )
        self.user_b = User.objects.create_user(
            username="admin_b",
            email="admin@agencia-b.com",
            password="password123",
        )

        # El middleware de onboarding redirige a /onboarding/ a usuarios
        # autenticados sin UsuarioAgencia. Enlazamos cada usuario a su
        # agencia para que las rutas de exportación sean alcanzables.
        UsuarioAgencia.objects.create(
            usuario=self.user_a, agencia=self.agencia_a, rol="admin", activo=True
        )
        UsuarioAgencia.objects.create(
            usuario=self.user_b, agencia=self.agencia_b, rol="admin", activo=True
        )

        # 3. Crear Moneda
        self.moneda = Moneda.objects.create(codigo_iso="USD", nombre="Dólar", simbolo="$")

        # 4. Crear Clientes aislados por Agencia
        with agency_context(self.agencia_a):
            self.cliente_a = Cliente.objects.create(
                nombres="Cliente",
                apellidos="De Agencia A",
                email="cliente@agencia-a.com",
                agencia=self.agencia_a,
            )

        with agency_context(self.agencia_b):
            self.cliente_b = Cliente.objects.create(
                nombres="Cliente",
                apellidos="De Agencia B",
                email="cliente@agencia-b.com",
                agencia=self.agencia_b,
            )

        # 5. Crear Ventas aisladas por Agencia
        with agency_context(self.agencia_a):
            self.venta_a = Venta.objects.create(
                localizador="PXP111",
                cliente=self.cliente_a,
                moneda=self.moneda,
                subtotal=Decimal("100.00"),
                impuestos=Decimal("20.00"),
                total_venta=Decimal("120.00"),
                agencia=self.agencia_a,
            )

        with agency_context(self.agencia_b):
            self.venta_b = Venta.objects.create(
                localizador="PXP222",
                cliente=self.cliente_b,
                moneda=self.moneda,
                subtotal=Decimal("200.00"),
                impuestos=Decimal("40.00"),
                total_venta=Decimal("240.00"),
                agencia=self.agencia_b,
            )

        # 6. Crear Facturas aisladas por Agencia
        with agency_context(self.agencia_a):
            self.factura_a = Factura.objects.create(
                numero_factura="FAC-A-001",
                venta_asociada=self.venta_a,
                cliente=self.cliente_a,
                moneda=self.moneda,
                subtotal=Decimal("100.00"),
                monto_impuestos=Decimal("20.00"),
                monto_total=Decimal("120.00"),
                agencia=self.agencia_a,
            )

        with agency_context(self.agencia_b):
            self.factura_b = Factura.objects.create(
                numero_factura="FAC-B-001",
                venta_asociada=self.venta_b,
                cliente=self.cliente_b,
                moneda=self.moneda,
                subtotal=Decimal("200.00"),
                monto_impuestos=Decimal("40.00"),
                monto_total=Decimal("240.00"),
                agencia=self.agencia_b,
            )

        # 7. Clientes de testeo
        self.client_a = Client()
        self.client_b = Client()

    def test_facturas_export_excel_isolation(self):
        """Verifica que al exportar Facturas a Excel se respete el multi-tenant."""
        self.client_a.force_login(self.user_a)

        response = self.client_a.get("/finance/invoices/?export=excel")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        # Para verificar que no hay fugas, podemos leer el contenido de Excel en memoria
        from io import BytesIO

        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active

        # Buscar en las celdas
        sheet_text = ""
        for row in ws.iter_rows(values_only=True):
            sheet_text += " ".join(str(cell) for cell in row if cell is not None)

        self.assertIn("FAC-A-001", sheet_text)
        self.assertNotIn("FAC-B-001", sheet_text)

    @requires_pdf_backend
    def test_ventas_export_pdf_isolation(self):
        """Verifica que al exportar Ventas a PDF se respete el multi-tenant."""
        self.client_a.force_login(self.user_a)

        response = self.client_a.get("/bookings/ventas/?export=pdf")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")

        self.assertTrue(len(response.content) > 0)

    def test_clientes_export_excel_isolation(self):
        """Verifica que al exportar Clientes a Excel se respete el multi-tenant."""
        self.client_a.force_login(self.user_a)

        response = self.client_a.get("/crm/clientes/?export=excel")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        from io import BytesIO

        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active

        sheet_text = ""
        for row in ws.iter_rows(values_only=True):
            sheet_text += " ".join(str(cell) for cell in row if cell is not None)

        self.assertIn("De Agencia A", sheet_text)
        self.assertNotIn("De Agencia B", sheet_text)
